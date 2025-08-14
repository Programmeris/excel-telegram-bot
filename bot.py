import os
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from openpyxl import load_workbook

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
XLSX_FILE_PATH = os.getenv("XLSX_FILE_PATH")

async def find(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message and update.message.text:
        workbook = load_workbook(XLSX_FILE_PATH)
        sheet = workbook.active
        headers = get_headers(sheet)
        keywords_column = headers.get('Keywords')
        links_column = headers.get('Link')
        response_message = ''

        if keywords_column and links_column:
            for row in sheet.iter_rows():
                parsed_message = update.message.text.replace("/find", "").strip()
                keywords = str(row[keywords_column - 1].value).replace(" ", "").split(";")
                keyword_match = any(element in parsed_message for element in keywords)
                if keyword_match:
                    response_message += row[links_column - 1].value
                
        await update.message.reply_text(f"{response_message}")

def get_headers(sheet: any ) -> dict:
    headers = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        headers[cell.value] = col_idx
    
    return headers

def main():
    try:
        app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()
        app.add_handler(CommandHandler("find", find))
        app.run_polling()
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")

if __name__ == "__main__":
    main()
